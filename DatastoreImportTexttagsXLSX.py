#!/usr/bin/env python3
from google.cloud import datastore
from collections import namedtuple
from openpyxl import load_workbook
import time
from XLIFFToXLSX import process_xliff_soup
from XLIFFReader import findXLIFFFiles, parse_xliff_file
from concurrent.futures import ThreadPoolExecutor
from ansicolor import black
from AutoTranslationIndexer import TextTagIndexer
from DatastoreUtils import DatastoreChunkClient

client = datastore.Client(project="watts-198422")
executor = ThreadPoolExecutor(512)
chunkClient = DatastoreChunkClient(client, executor)

def update_texttag(lang, engl, transl):
    key = client.key('Texttag', engl, namespace=lang)
    obj = client.get(key)
    if obj is not None:
        obj.update({"translated": transl, "approved_in_ui": True})
        client.put(obj)
        print(obj)


def read_texttags_xlsx(filename):
    wb = load_workbook(filename=filename)
    sheet = wb[wb.sheetnames[0]]
    tmap = {}
    for row in sheet.rows:
        engl = row[0]
        transl = row[1]
        if engl == "Source":
            continue
        if transl.value is None:
            continue
        yield str(engl.value), str(transl.value)

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('lang', help='The crowdin lang code')
    parser.add_argument('xlsx', help='The xlsx to import')
    args = parser.parse_args()

    for engl, transl in read_texttags_xlsx(args.xlsx):
        executor.submit(update_texttag, args.lang, engl, transl)

