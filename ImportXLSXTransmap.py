#!/usr/bin/env python3
from openpyxl import load_workbook
import json
import argparse
import os
from ansicolor import red, blue, green
from AutoTranslateCommon import transmap_filename

def get_transmap(filename):
    wb = load_workbook(filename=filename)
    sheet = wb[wb.sheetnames[0]]
    tmap = {}
    for row in sheet.rows:
        engl = row[0]
        transl = row[1]
        if transl.value is None:
            continue
        print(engl.value,"==>",transl.value)

        tmap[str(engl.value)] = str(transl.value)
    return [{"english": engl, "translated": transl} for engl,transl in tmap.items()]

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('-l','--language', help='The language')
    parser.add_argument('iftags', help='The IF tags XLSX file')
    parser.add_argument('texttags', help='The text tags XLSX file')
    args = parser.parse_args()

    if os.path.isfile(args.iftags):
        iftags = get_transmap(args.iftags)
        print("Found {} iftags".format(len(iftags)))
        
        iftagsFile = transmap_filename(args.language, "ifpatterns")
        print("Exporting IF tags to {}".format(iftagsFile))
        
        with open(iftagsFile, "w") as outfile:
            json.dump(iftags, outfile, indent=4, sort_keys=True)
    else:
        print(red("IFTags file doesn't exist, skipping...", bold=True))

    if os.path.isfile(args.texttags):
        texttags = get_transmap(args.texttags)
        print("Found {} text tags".format(len(texttags)))

        texttagsFile = transmap_filename(args.language, "texttags")
        print("Exporting text tags to {}".format(texttagsFile))

        with open(texttagsFile, "w") as outfile:
            json.dump(texttags, outfile, indent=4, sort_keys=True)
    else:
        print(red("Texttags file doesn't exist, skipping...", bold=True))
    
